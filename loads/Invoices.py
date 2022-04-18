import os
from io import BytesIO
from django.conf import settings
import openpyxl as xl
from tempfile import NamedTemporaryFile
import boto3


s3 = boto3.client('s3',  aws_access_key_id=settings.AWS_ACCESS_KEY_ID, aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY, region_name='us-east-1')


# This script creates a basic excel invoice. Class InvoiceItems and Customers store relevant invoice data based on user input
# This data is then refernced by the Invoice class in order to generate the invoice. 

class InvoiceItem():
    def __init__(self, description, qty, unit_price):
        self.description = description
        self.qty = qty
        self.unit_price = unit_price


class Customers():
    def __init__(self, name, street, city, state, zip, phone, email):
        self.name = name
        self.street = street
        self.city = city
        self.state = state
        self.zip = zip
        self.phone = phone
        self.email = email


class Invoices():
    def __init__(self, customer: Customers):
        self.item_list = []
        self.customer = customer
        

    # Add separate items to be invoiced on the same invoice
    def add_item(self, description, unit_price, qty):
        self.item_list.append(InvoiceItem(description, qty, unit_price))

    # Determine where the invoice should be saved in the aws s3 bucket. Creates new directory based on user input
    # if one does not already exist
    def create_directory(self, directory):
        parent_dir = "loads/"
        path = parent_dir + directory
        return path


        # parent_dir = "/Users/sm/Desktop/Symport"
        # path = os.path.join(parent_dir, directory)
        # if os.path.isdir(path):
        #     return path
        # os.makedirs(path)
        # return path





    # this function generates the final invoice using information passed through from the functions above. 
    # The function uses a pre generated excel template stored in the parent directory. 
    def create_invoice(self, directory, file_name):
        
        path = self.create_directory(directory)

        s3_data = s3.get_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key='Symport_Invoice_Template.xlsx')
        contents = s3_data['Body'].read()
        wb = xl.load_workbook(filename=(BytesIO(contents)), data_only=True)
        sheet = wb["Invoice"]
        

        #Code for working with files on local drive
        # template_path = "/Users/sm/Desktop/Symport/Invoices"
        # template_file_name = "/Symport_Invoice_Template.xlsx"
        # wb = xl.load_workbook(template_path + template_file_name)
        # sheet = wb["Invoice"]


        sheet.cell(row=8, column=1).value = self.customer.name
        sheet.cell(row=9, column=1).value = self.customer.street
        sheet.cell(row=10, column=1).value = self.customer.city + " " + self.customer.state + " " + self.customer.zip
        sheet.cell(row=11, column=1).value = self.customer.phone
        sheet.cell(row=12, column=1).value = self.customer.email

        row_increment = 16

        # Add each item to be invoiced to the invoice
        for item in self.item_list:
            sheet.cell(row=row_increment, column=1).value = item.description
            sheet.cell(row=row_increment, column=6).value = item.qty
            sheet.cell(row=row_increment, column=7).value = item.unit_price
            row_increment += 1

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            output = BytesIO(tmp.read())
            s3.put_object(Body=output, Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=(path + "/" + file_name + ".xlsx"))
        
        
        # wb.save(path + "/" + file_name + ".xlsx")






# customer = Customers("detailed", "9219 Knight Ave", "Des Plaines", "IL", "60016", "77334994", "smotlani@yahoo.com")
# invoice = Invoices(customer)
# invoice.add_item("load12345", 1, 350)
# invoice.add_item("detention", 2, 50)
# invoice.create_invoice('AGT', 'a')